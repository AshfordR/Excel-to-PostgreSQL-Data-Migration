import openpyxl
import psycopg2
from psycopg2 import IntegrityError  # Import the IntegrityError class

# Prompt the user for the file path
data = input("Enter the file path: ")

try:
    # Load the Excel workbook
    excel_data = openpyxl.load_workbook(data)
except FileNotFoundError:
    print(f"File '{data}' not found. Please provide a valid file path.")
    exit(1)

# Prompt the user for the sheet name
sheet_name = input("Enter the sheet name: ")

# Check if the sheet name exists in the workbook
if sheet_name not in excel_data.sheetnames:
    print(f"Sheet '{sheet_name}' not found in the workbook.")
    exit(1)

# Access the sheet by its name
sheet = excel_data[sheet_name]
print()

# Prompt the user for database connection parameters
db_host = input("Enter the database host: ")
db_name = input("Enter the database name: ")
db_user = input("Enter the database user: ")
db_password = input("Enter the database password: ")

# Connect to PostgreSQL using the user-provided parameters
try:
    conn = psycopg2.connect(
        host=db_host,
        database=db_name,
        user=db_user,
        password=db_password
    )
except psycopg2.Error as e:
    print(f"Error connecting to the database: {e}")
    exit(1)

cur = conn.cursor()

# Create an empty list to store the extracted data
extracted_data = []

# Get the header row directly
header_row = list(sheet.iter_rows(min_row=9, max_row=9, min_col=3, max_col=20, values_only=True))[0]
print(header_row)


# Fetch existing group names from the database
cur.execute("SELECT name FROM auth_group")
existing_group_names = [row[0] for row in cur.fetchall()]

# Check if any header in header_row is not present in existing_group_names
missing_groups = [header for header in header_row if header not in existing_group_names]

# Insert missing groups into auth_group table
for missing_group in missing_groups:
    try:
        cur.execute("INSERT INTO auth_group (name) VALUES (%s)", (missing_group,))
        conn.commit()
        print(f"Added missing group: {missing_group} to auth_group table")
    except IntegrityError as e:
        # Handle IntegrityError as needed (e.g., if the group name already exists)
        print(f"Error adding group: {missing_group} - {e}")
    except Exception as e:
        # Handle other exceptions as needed
        print(f"An error occurred while adding group: {missing_group} - {e}")

# Iterate through the rows to extract permission names with 'X'
for row in sheet.iter_rows(min_row=9, max_row=279, min_col=3, max_col=20, values_only=True):
    permission = row[0]

    # Extract group names with 'X' from the corresponding columns
    group_names = [header_row[i] for i, cell_value in enumerate(row[1:], start=1) if cell_value == 'X']

    extracted_data.append((permission, group_names))


# ... (Continue with matching permission names and inserting into auth_group_permissions)
for permission, group_names in extracted_data:
    print(f"Permission: {permission}")
    print(f"Group Names with 'X': {group_names}")
    print()

# Begin a transaction
#conn.autocommit = False
group_id = None  # Initialize group_id with a default value

# Match permission names and group names to permission IDs and group IDs
for permission, group_names in extracted_data:
    for group_name in group_names:
        try:
            # Fetch permission ID based on permission name
            cur.execute("SELECT id FROM auth_permission WHERE name = %s", (permission,))
            permission_id_result = cur.fetchone()

            if permission_id_result is not None:
                permission_id = permission_id_result[0]
            else:
                # Handle the case where no permission ID was found
                print(f"Permission name '{permission}' not found in auth_permission.")
                continue

            # Fetch group ID based on group name
            cur.execute("SELECT id FROM auth_group WHERE name = %s", (group_name,))
            group_id_result = cur.fetchone()

            if group_id_result is not None:
                group_id = group_id_result[0]
            else:
                # Handle the case where no group ID was found
                print(f"Group name '{group_name}' not found in auth_group.")
                continue

            # Insert into auth_group_permission table
            cur.execute("INSERT INTO auth_group_permissions (group_id, permission_id) VALUES (%s, %s)", (group_id, permission_id))

        except IntegrityError as e:
            # Handle the UniqueViolation exception (duplicate key) and rollback
            if "duplicate key value" in str(e):
                conn.rollback()
            else:
                # Handle other IntegrityError exceptions as needed
                print("An IntegrityError occurred:", e)
        finally:
            # Commit the transaction (whether it was inserted or rolled back)
            conn.commit()

# Close database connection
cur.close()
conn.close()
